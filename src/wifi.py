import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the path to the input and output files
input_file = 'path/to/your/input_file.xlsx'  # Original file
output_file = 'path/to/your/output_file_with_analysis.xlsx'  # Final output file

# Load the workbook and read the data, skipping any metadata rows if needed
book = load_workbook(input_file)
data = pd.read_excel(input_file, skiprows=2)

# Adjust column names based on the file structure, as observed
data.columns = ['DateTime', 'SupplyTemp', 'ReturnTemp', 'Mode', 'Request', 'State', 'Status']

# Split 'DateTime' into separate 'Date' and 'Time' columns
data[['Date', 'Time']] = data['DateTime'].astype(str).str.split(' ', expand=True)
data.drop(columns=['DateTime'], inplace=True)

# Filter rows where Status is "Test Run" or similar
testing_data = data[data['Status'].str.contains('Test', na=False, case=False)].copy()
testing_data['Switch'] = testing_data['State'].ne(testing_data['State'].shift()).astype(int)
switch_data = testing_data[testing_data['Switch'] == 1].copy()

# Calculate deviations in decimal hours and mm:ss format
switch_data['Datetime'] = pd.to_datetime(switch_data['Date'] + ' ' + switch_data['Time'], dayfirst=True, errors='coerce')
switch_data.dropna(subset=['Datetime'], inplace=True)
switch_data['Time_Diff'] = switch_data['Datetime'].diff().dt.total_seconds() / 3600  # time difference in hours
switch_data['Decimal_Deviation'] = switch_data['Time_Diff'] - 2  # deviation from a 2-hour baseline
switch_data['Deviation_Min_Sec'] = (switch_data['Decimal_Deviation'] * 3600).apply(
    lambda x: f"{int(abs(x) // 60)}:{int(abs(x) % 60):02}" if pd.notnull(x) else None
)

# Create a new sheet in the original workbook for the deviation analysis
deviation_analysis_sheet = book.create_sheet(title="Deviation Analysis")

# Write the analysis data to the new sheet
for r_idx, row in enumerate(dataframe_to_rows(switch_data[['Datetime', 'State', 'Decimal_Deviation', 'Deviation_Min_Sec']], index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        deviation_analysis_sheet.cell(row=r_idx, column=c_idx, value=value)

# Define fills for conditional formatting
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply conditional formatting: yellow for deviations between 2 and 5 minutes, red for over 5 minutes
deviation_analysis_sheet.conditional_formatting.add(
    'C2:C1000',
    CellIsRule(operator='between', formula=['0.0333', '0.0833'], fill=yellow_fill)
)
deviation_analysis_sheet.conditional_formatting.add(
    'C2:C1000',
    CellIsRule(operator='greaterThan', formula=['0.0833'], fill=red_fill)
)

# Save the workbook with the new sheet and formatting
book.save(output_file)
print(f"Processed and saved analysis as a new sheet in {output_file}")
