import os
import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def main():
    parser = argparse.ArgumentParser(description="Process an Excel file and generate output with deviation analysis.")
    parser.add_argument("--input-file", required=True, help="Path to the input Excel file")
    args = parser.parse_args()

    input_file = args.input_file
    if not input_file.endswith('.xlsx'):
        print("The input file must be an .xlsx file.")
        return

    # Define output file path
    output_file = os.path.splitext(input_file)[0] + "_OUTPUT.xlsx"
    print(f"Processing file: {input_file}")

    # Load the workbook and save it as the output file
    print("Loading workbook and saving initial structure...")
    book = load_workbook(input_file)
    book.save(output_file)

    # Re-load the output file for adding a new sheet
    print("Re-loading saved workbook...")
    book = load_workbook(output_file)
    data = pd.read_excel(input_file, skiprows=2)

    # Remove rows without dates and filter relevant data
    print("Cleaning data and filtering relevant rows...")
    data.columns = ['Date', 'Time', 'SupplyTemp', 'ReturnTemp', 'Mode', 'Request', 'State', 'Status']
    data = data[data['Date'].str.match(r'^\d{2}/\d{2}/\d{4}$', na=False)].copy()
    data.reset_index(drop=True, inplace=True)

    # Filter rows for "testing" and switching states
    print("Filtering for 'testing' rows and identifying state switches...")
    testing_data = data[data['Status'].str.lower() == 'testing'].copy()
    testing_data['Switch'] = testing_data['State'].ne(testing_data['State'].shift()).astype(int)
    switch_data = testing_data[testing_data['Switch'] == 1].copy()

    # Calculate the time deviation in both decimal hours and mm:ss format
    print("Calculating time deviations and formatting data...")
    switch_data['Datetime'] = pd.to_datetime(
        switch_data['Date'].astype(str) + ' ' + switch_data['Time'].astype(str),
        dayfirst=True, errors='coerce'
    )
    switch_data.dropna(subset=['Datetime'], inplace=True)
    switch_data['Time_Diff'] = switch_data['Datetime'].diff().dt.total_seconds() / 3600  # in hours
    switch_data['Decimal_Deviation'] = switch_data['Time_Diff'] - 2  # 2 hours as baseline

    # Convert deviation in hours to mm:ss format for clarity
    switch_data['Deviation_Min_Sec'] = (switch_data['Decimal_Deviation'] * 3600).apply(
        lambda x: f"{int(abs(x) // 60)}:{int(abs(x) % 60):02}" if pd.notnull(x) else None
    )

    # Write deviation analysis to a new sheet
    print("Writing deviation analysis to a new sheet...")
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        writer.book = book  # Use the reloaded workbook
        switch_data[['Datetime', 'State', 'Decimal_Deviation', 'Deviation_Min_Sec']].to_excel(
            writer, sheet_name='Deviation Analysis', index=False
        )

    # Apply conditional formatting to the new sheet
    print("Applying conditional formatting...")
    deviation_analysis_sheet = book['Deviation Analysis']
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Yellow: for deviations over 2 minutes (0.0333 hours) but less than 5 minutes (0.0833 hours)
    deviation_analysis_sheet.conditional_formatting.add(
        'C2:C1000',
        CellIsRule(operator='between', formula=['0.0333', '0.0833'], fill=yellow_fill)
    )

    # Red: for deviations over 5 minutes (0.0833 hours)
    deviation_analysis_sheet.conditional_formatting.add(
        'C2:C1000',
        CellIsRule(operator='greaterThan', formula=['0.0833'], fill=red_fill)
    )

    # Save the workbook with updated sheet and formatting
    book.save(output_file)
    print(f"Processing complete. Output saved to {output_file}")

if __name__ == "__main__":
    main()
